import openpyxl
import os

def write_to_excel(invoice_data, excel_path=R'C:\Users\admin\Desktop\发票信息.xlsx', sheet_name='发票列表'):
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        sheet.append(['发票号码', '发票日期', '销售方', '金额', '税率', '税额', '价税合计'])
    else:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]

    # 避免重复写入
    existing_invoices = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
    if invoice_data['发票号码'] in existing_invoices:
        print(f"发票 {invoice_data['发票号码']} 已存在，跳过写入。")
        return

    sheet.append([
        invoice_data['发票号码'],
        invoice_data['发票日期'],
        invoice_data['销售方'],
        invoice_data['金额'],
        invoice_data['税率'],
        invoice_data['税额'],
        invoice_data['价税合计']
    ])
    wb.save(excel_path)
    print(f"发票 {invoice_data['发票号码']} 写入成功！")
